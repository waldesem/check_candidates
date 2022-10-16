from tkinter import filedialog, messagebox
from datetime import date
import check_info
import openpyxl
import getpass
import os
import sqlite3

def db_add():
    try:
        sqlite_connection = sqlite3.connect('/home/semenenko/MyProjects/Python/Staff_check/DB_check/candidates_db.db')
        cursor_obj = sqlite_connection.cursor()
        with sqlite_connection:
            cursor_obj.execute("INSERT INTO candidates(staff, department, full_name, last_name, birthday, birth_place, country, serie_passport, number_passport, date_given, snils, inn, reg_address, live_address, phone, email, education, first_time_work, first_place_work, check_first_place, second_time_work, second_place_work, check_second_place, third_time_work, third_place_work, check_third_place, check_passport, check_selfwork, check_inn, check_debt, check_bancrupcy, check_bki, check_affilate, check_disqual, check_db, check_internet, check_cronos, check_cross, resume, date_check, officer, url) VALUES ("+"'"+anketa_values[0]+"', "+"'"+anketa_values[1]+"', "+"'"+anketa_values[2]+"', "+"'"+anketa_values[3]+"', "+"'"+anketa_values[4]+"', "+"'"+anketa_values[5]+"', "+"'"+anketa_values[6]+"', "+"'"+anketa_values[7]+"', "+"'"+anketa_values[8]+"', "+"'"+anketa_values[9]+"', "+"'"+anketa_values[10]+"', "+"'"+anketa_values[11]+"', "+"'"+anketa_values[12]+"', "+"'"+anketa_values[13]+"', "+"'"+anketa_values[14]+"', "+"'"+anketa_values[15]+"', "+"'"+anketa_values[16]+"', "+"'"+anketa_values[17]+"', "+"'"+anketa_values[18]+"', 'NULL', "+"'"+anketa_values[19]+"', "+"'"+anketa_values[20]+"', 'NULL', "+"'"+anketa_values[21]+"', "+"'"+anketa_values[22]+"', 'NULL', "+"'"+check_info.response_check[4]+"', "+"'"+check_info.response_check[0]+"', "+"'"+check_info.response_check[1]+"', 'NULL', 'NULL', 'NULL', 'NULL', "+"'"+check_info.response_check[3]+"', "+"'"+check_info.response_check[2]+"', 'NULL', 'NULL', 'NULL', 'NULL', "+"'"+date.today().strftime('%d.%m.%Y')+"', "+"'"+str(getpass.getuser())+"', 'NULL')")
            
    except sqlite3.Error as error:
        print('Ошибка при подключении к sqlite', error)
    finally:
        if sqlite_connection:
            sqlite_connection.close()

def upload():
    file = filedialog.askopenfilename(filetypes=[("Excel files", ".xlsx .xlsm")])
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.worksheets[0]
    global anketa_values
    anketa_values = [str(ws['C3'].value), str(ws['D3'].value), str(ws['K3'].value), str(ws['S3'].value),
               str(ws['L3'].value), str(ws['M3'].value), str(ws['T3'].value), str(ws['P3'].value),
               str(ws['Q3'].value), str(ws['R3'].value), str(ws['U3'].value), str(ws['V3'].value),
               str(ws['N3'].value), str(ws['O3'].value), str(ws['Y3'].value), str(ws['Z3'].value),
               str(ws['X3'].value), str(ws['AA3'].value), str(ws['AB3'].value), str(ws['AA4'].value),
               str(ws['AB4'].value), str(ws['AA5'].value), str(ws['AB5'].value)]
    wb.close()

    os.makedirs('/home/semenenko/MyProjects/' + anketa_values[2], exist_ok=True)
    wb.save('/home/semenenko/MyProjects/' + anketa_values[2] + '/' + 'Анкета ' +
            anketa_values[2] + '-' + anketa_values[4] + '.xlsx')

    return anketa_values

def download_file():
    wb = openpyxl.load_workbook(r'/home/semenenko/MyProjects/Python/Staff_check/DB_check/Заключение.xlsx')
    ws = wb['Заключение']
    # анкетные данные
    ws['C4'] = anketa_values[0]
    ws['C5'] = anketa_values[1]
    ws['C6'] = anketa_values[2]
    ws['C7'] = anketa_values[4]
    # 'Результаты проверки по местам работы'
    ws['A9'] = anketa_values[-6]
    ws['C9'] = anketa_values[-5]
    ws['A10'] = anketa_values[-4]
    ws['C10'] = anketa_values[-3]
    ws['A11'] = anketa_values[-2]
    ws['C11'] = anketa_values[-1]
    # проверка паспорт
    ws['C13'] = check_info.response_check[4]
    # 'Проверка статуса самозанятого'
    ws['C14'] = check_info.response_check[0]
    # 'Проверка ИНН'
    ws['C15'] = check_info.response_check[1]
    # 'Проверка по списку кандидатов'
    ws['C21'] = check_info.response_check[2]
    # 'Проверка по списку дисквалифицированных лиц'
    ws['C20'] = check_info.response_check[3]
    # 'Дата проверки'
    ws['C26'] = date.today().strftime('%d.%m.%Y')
    # имя пользователя
    ws['C27'] = getpass.getuser()

    # создаем папку и сохраняем файл с новым именем
    wb.save('/home/semenenko/MyProjects/' + anketa_values[2] + '/' + 'Заключение ' +
            anketa_values[2] + '-' + anketa_values[4] + '.xlsx')
    
    db_add()

    messagebox.showinfo(title="Результат проверки", message="Создана папка c файлами проверки. Добавлена запись в БД")
