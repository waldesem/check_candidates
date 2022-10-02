import datetime as dt
import httpx
import openpyxl
import sqlite3

import menu_upload

response_check = None
response_candidates = None
response_disqual = None


def check():
    # проверка по списку самозанятых
    def check_npd_status(inn: str, req_date: dt.date = None) -> dict:
        req_date = dt.date.today().isoformat()
        url = "https://statusnpd.nalog.ru/api/v1/tracker/taxpayer_status"
        data = {
            "inn": inn,
            "requestDate": req_date,
        }
        resp = httpx.post(url=url, json=data)
        return resp.json()

    # проверка инн
    def suggest(surname: str, name: str, patronymic: str, birthdate: str,
                doctype: str, docnumber: str, docdate: str) -> dict:
        url = "https://service.nalog.ru/inn-proc.do"
        data = {
            "fam": surname,
            "nam": name,
            "otch": patronymic,
            "bdate": birthdate,
            "bplace": "",
            "doctype": doctype,
            "docno": docnumber,
            "docdt": docdate,
            "c": "innMy",
            "captcha": "",
            "captchaToken": "",
        }
        resp = httpx.post(url=url, data=data)
        resp.raise_for_status()
        return resp.json()

    # проверка по списку кандидатов
    def candidates():
        # file = r'\\cronosx1\New folder\УВБ\Отдел корпоративной защиты\Кандидаты\Кандидаты.xlsm'
        global response_candidates
        file = r'/home/semenenko/MyProjects/Python/Staff_check/DB_check/Кандидаты.xlsm'
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        col_range = ws['B1':'B30000']
        flag = False
        for cell in col_range:
            for c in cell:
                if str(c.value).strip() == menu_upload.anketa_values[2].strip():
                    row_num = c.row
                    birthday = ws['C' + str(row_num)].value
                    birthday = dt.datetime.strftime(birthday, '%d.%m.%Y')
                    if birthday == menu_upload.anketa_values[4]:
                        response_candidates = 'Найдены полные совпадения в списке Кандидатов по ФИО и дате рождения'
                        flag = True
                        break
                    else:
                        continue
                else:
                    response_candidates = 'Не найдены совпадения в списке Кандидатов (возможно совпадение по ФИО)'
            if flag:
                break
        wb.close()
        return response_candidates

    # проверка по списку дисквалифицированных
    def disqualified():
        fio = str(menu_upload.anketa_values[2]).upper()
        dr = str(menu_upload.anketa_values[3])
        global response_disqual
        try:
            sqlite_connection = sqlite3.connect('/home/semenenko/MyProjects/Python/Staff_check/DB_check/disqual_db.db')
            cursor_obj = sqlite_connection.cursor()
            cursor_obj.execute("SELECT * FROM disqual_fns WHERE G2 like " + "'" + fio + "'" 
                                + ' and G3 like ' + "'" + dr + "'")
            record = cursor_obj.fetchall()
            if len(record) != 0:
                response_disqual = 'Найден в списке дисквалифицированных лиц'
            else:
                response_disqual = 'Отсутствует в списке дисквалифицированных лиц'
            cursor_obj.close()
        except sqlite3.Error as error:
            print("Ошибка при подключении к sqlite", error)
        finally:
            if sqlite_connection:
                sqlite_connection.close()
        return response_disqual

    # проверка по списку паспортов в БД
    def passport():
        try:
            sqlite_connection = sqlite3.connect('/home/semenenko/Загрузки/passportDB.db')
            cursor_obj = sqlite_connection.cursor()
            cursor_obj.execute('SELECT * FROM list_of_expired_passports WHERE PASSP_SERIES = ' + str(
                menu_upload.anketa_values[7]) + ' and PASSP_NUMBER = ' + str(menu_upload.anketa_values[8]))
            record = cursor_obj.fetchall()
            if len(record) != 0:
                passport_response = "Найден в списке недействительных паспортов"
            else:
                passport_response = 'В списке недействительных паспортов не найден'
            cursor_obj.close()
        except sqlite3.Error as error:
            print("Ошибка при подключении к sqlite", error)
        finally:
            if sqlite_connection:
                sqlite_connection.close()
        return passport_response

    # собираем проверенные данные
    global response_check
    response_check = []
    # проверка самозанятого
    npd_response = check_npd_status(menu_upload.anketa_values[11])
    response_check.append(npd_response.get('message'))
    # проверка инн
    document = {'passport_foreign': '10', 'residence_permit': '12', 'passport_russia': '21'}
    inn_response = suggest(
        surname=menu_upload.anketa_values[2].split()[0],
        name=menu_upload.anketa_values[2].split()[1],
        patronymic=menu_upload.anketa_values[2].split()[2],
        birthdate=menu_upload.anketa_values[4],
        doctype=document.get('passport_russia'),
        docnumber=menu_upload.anketa_values[7][:2] + ' ' + menu_upload.anketa_values[7][2:] + ' ' + menu_upload.anketa_values[8],
        docdate=menu_upload.anketa_values[9],
    )
    # if __name__ == "__main__":
    response_check.append(inn_response.get('inn'))
    # проверка по списку Кандидатов
    candidates_response = candidates()
    response_check.append(candidates_response)
    # проверка по списку Кандидатов
    response_disqual = disqualified()
    response_check.append(response_disqual)
    # проверка по списку паспортов
    passport_response = passport()
    response_check.append(passport_response)
    return response_check
