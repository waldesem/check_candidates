from tkinter import filedialog, messagebox
from datetime import date
from database_app import response_db
import requests
import httpx
import openpyxl
import getpass
import os

# загрузить анкету в формате Excel
def upload():
    file = filedialog.askopenfilename(filetypes=[("Excel files", ".xlsx .xlsm")])
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.worksheets[0]
    global anketa_values
    anketa_values = [str(i) for i in [ws['C3'].value, ws['D3'].value, ws['K3'].value, ws['S3'].value, ws['L3'].value, ws['M3'].value, ws['T3'].value, ws['P3'].value, ws['Q3'].value, ws['R3'].value, ws['U3'].value, ws['V3'].value, ws['N3'].value, ws['O3'].value, ws['Y3'].value, ws['Z3'].value, ws['X3'].value, ws['AA3'].value, ws['AB3'].value, ws['AA4'].value, ws['AB4'].value, ws['AA5'].value, ws['AB5'].value]]
    wb.close()
    # создаем папку для материалов проверки
    os.makedirs('/home/semenenko/MyProjects/' + anketa_values[2], exist_ok=True)
    wb.save('/home/semenenko/MyProjects/' + anketa_values[2] + '/' + 'Анкета ' + anketa_values[2] + '-' + anketa_values[4] + '.xlsx')
    return anketa_values

# выгрузить анкету и результаты проверки в БД
def db_add():
    db = '/home/semenenko/MyProjects/Python/Staff_check/DB_check/candidates_db.db'
    query = ("INSERT INTO candidates(staff, department, full_name, last_name, birthday, birth_place, country, serie_passport, number_passport, date_given, snils, inn, reg_address, live_address, phone, email, education, first_time_work, first_place_work, check_first_place, second_time_work, second_place_work, check_second_place, third_time_work, third_place_work, check_third_place, check_passport, check_terror, check_selfwork, check_inn, check_debt, check_bancrupcy, check_bki, check_affilate, check_disqual, check_db, check_internet, check_cronos, check_cross, resume, date_check, officer, url) VALUES ("+"'"+anketa_values[0]+"', "+"'"+anketa_values[1]+"', "+"'"+anketa_values[2]+"', "+"'"+anketa_values[3]+"', "+"'"+anketa_values[4]+"', "+"'"+anketa_values[5]+"', "+"'"+anketa_values[6]+"', "+"'"+anketa_values[7]+"', "+"'"+anketa_values[8]+"', "+"'"+anketa_values[9]+"', "+"'"+anketa_values[10]+"', "+"'"+anketa_values[11]+"', "+"'"+anketa_values[12]+"', "+"'"+anketa_values[13]+"', "+"'"+anketa_values[14]+"', "+"'"+anketa_values[15]+"', "+"'"+anketa_values[16]+"', "+"'"+anketa_values[17]+"', "+"'"+anketa_values[18]+"', 'NULL', "+"'"+anketa_values[19]+"', "+"'"+anketa_values[20]+"', 'NULL', "+"'"+anketa_values[21]+"', "+"'"+anketa_values[22]+"', 'NULL', "+"'"+response_check[4]+"', 'NULL', "+"'"+response_check[0]+"', "+"'"+response_check[1]+"', 'NULL', 'NULL', 'NULL', 'NULL', "+"'"+response_check[3]+"', "+"'"+response_check[2]+"', 'NULL', 'NULL', 'NULL', 'NULL', "+"'"+date.today().strftime('%d.%m.%Y')+"', "+"'"+str(getpass.getuser())+"', 'NULL')")
    add_db = response_db(db, query)
    return add_db

# переменная со списком для резульатов проверки
response_check = []

# проверка кандидата по источникам информации
def check():
    # проверка инн
    url_inn = "https://service.nalog.ru/inn-proc.do"
    data_inn = {
        "fam": anketa_values[2].split()[0],
        "nam": anketa_values[2].split()[1],
        "otch": anketa_values[2].split()[2],
        "bdate": anketa_values[4],
        "bplace": "",
        "doctype": '21',
        "docno": anketa_values[7][:2] + ' ' + anketa_values[7][2:] + ' ' + anketa_values[8],
        "docdt": anketa_values[9],
        "c": "innMy",
        "captcha": "",
        "captchaToken": "",
        }
    try:
        inn_resp = requests.post(url=url_inn, data=data_inn)
        inn_resp.raise_for_status()
        inn_response = inn_resp.json()
    except ConnectionError:
        messagebox.showwarning(title='Ошибка соединения', message='Попробуйте запустить проверку еще раз')

    # проверка самозанятого
    if anketa_values[11] == 'None' or anketa_values[11] == inn_response.get('inn'):
        url_npd = "https://statusnpd.nalog.ru/api/v1/tracker/taxpayer_status"
        data_npd = {
        "inn": inn_response.get('inn'), 
        "requestDate": date.today().isoformat(),
        }
        try:
            npd_resp = httpx.post(url=url_npd, json=data_npd)
            npd_response = npd_resp.json()
        except ConnectionError:
            messagebox.showwarning(title='Ошибка соединения', message='Попробуйте запустить проверку еще раз')
    else:
        npd_response = 'ИНН в анкете не соответствует даннымм ответа ФНС'

    # проверка по собственной Базе данных
    connect = '/home/semenenko/MyProjects/Python/Staff_check/DB_check/candidates_db.db'
    query = "SELECT * FROM candidates WHERE full_name like " + "'" + anketa_values[2] + "'" + ' and birthday like ' + "'" + anketa_values[4] + "'"
    if response_db(connect, query):
        candidates_response = 'Найдены полные совпадения в Базе данных по ФИО и дате рождения'
    else:
        candidates_response = 'Не найдены совпадения в списке Базе данных (возможно совпадение по ФИО)'
    
    # проверка по списку Дисквалификации
    connect = '/home/semenenko/MyProjects/Python/Staff_check/DB_check/disqual_db.db'
    query = "SELECT * FROM disqual_fns WHERE G2 like " + "'" + anketa_values[2].upper() + "'" + ' and G3 like ' + "'" + anketa_values[3] + "'"
    if response_db(connect, query):
        response_disqual = 'Найден в списке дисквалифицированных лиц'
    else:
        response_disqual = 'Отсутствует в списке дисквалифицированных лиц'
    
    # проверка по списку Недействительных паспортов
    connect = '/home/semenenko/Загрузки/passportDB.db'
    query = 'SELECT * FROM list_of_expired_passports WHERE PASSP_SERIES = ' + anketa_values[7] + ' and PASSP_NUMBER = ' + anketa_values[8]
    if response_db(connect, query):
        passport_response = 'Найден в списке недействительных паспортов'
    else:
        passport_response = 'В списке недействительных паспортов не найден'
    
    try:
        response_check.extend((npd_response.get('message'), inn_response.get('inn'), candidates_response, response_disqual, passport_response))
        messagebox.showinfo(title='Успех', message='Проверка успешно окончена')
    except AttributeError as error:
        messagebox.showwarning(title='Ошибка', message=error+'Попробуйте еще раз')

    return [str(i) for i in response_check]
