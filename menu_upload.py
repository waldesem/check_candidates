from tkinter import filedialog
import openpyxl

anketa_values = None
anketa_dict = None
anketa_labeles = ['Должность', 'Подразделение', 'Фамилия Имя Отчество', 'Предыдущее ФИО', 'Дата рождения',
                    'Место рождения', 'Гражданство', 'Серия паспорта', 'Номер паспорта', 'Дата выдачи', 
                    'СНИЛС', 'ИНН', 'Адрес регистрации', 'Адрес проживания', 'Телефон', 'Электронная  почта', 
                    'Образование', 'Период работы', 'Место работы', 'Период работы', 'Место работы', 
                    'Период работы', 'Место работы']

def upload():
    file = filedialog.askopenfilename(filetypes=[("Excel files", ".xlsx .xlsm")])
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    global anketa_values, anketa_labeles, anketa_dict
    anketa_values = [ws['C3'].value, ws['D3'].value, ws['K3'].value, ws['S3'].value,
               ws['L3'].value, ws['M3'].value, ws['T3'].value, ws['P3'].value,
               ws['Q3'].value, ws['R3'].value, ws['U3'].value, ws['V3'].value,
               ws['N3'].value, ws['O3'].value, ws['Y3'].value, ws['Z3'].value,
               ws['X3'].value, ws['AA3'].value, ws['AB3'].value, ws['AA4'].value,
               ws['AB4'].value, ws['AA5'].value, ws['AB5'].value]
    
    anketa_dict = dict(zip(anketa_labeles, anketa_values))
    wb.close()
    return anketa_values
