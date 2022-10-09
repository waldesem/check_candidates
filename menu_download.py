from tkinter import messagebox as mb
from datetime import date
import openpyxl
import menu_upload
import check_main
import getpass
import sqlite3
import os

# записываем результаты в базу данных
# def db_add():
#     try:
#         sqlite_connection = sqlite3.connect('/home/semenenko/MyProjects/Python/Staff_check/DB_check/kandidates.db')
#         cursor_obj = sqlite_connection.cursor()
#         cursor_obj.execute("SELECT * FROM Candidates WHERE ФИО like " + "'" + fio_search.get() + 
#                             "'" + ' and Датарождения like ' + "'" + dr_search.get() + "'")
#         record_db = cursor_obj.fetchall()
#         try:
#             if record_db[0][1] == fio_search.get() and record_db[0][2] == dr_search.get():
#                 for i in range(3):
#                     create_labels(tab_db,txt_db[i], 40, 0, i+2)
#                     create_labels(tab_db,record_db[0][i], 60, 1, i+2)
#         except IndexError:
#             mb.showinfo(title="Внимание", message="Запись не найдена")
#         cursor_obj.close()
#     except sqlite3.Error as error:
#         mb.showinfo(title = "Внимание", message = "Ошибка при подключении к sqlite" + error)
#     finally:
#         if sqlite_connection:
#             sqlite_connection.close()
#     return record_db

# today_date = date.today().strftime('%d.%m.%Y')
def download_file():
    wb = openpyxl.load_workbook(r'/home/semenenko/MyProjects/Python/Staff_check/DB_check/Заключение.xlsm')
    ws = wb['Заключение']
    # анкетные данные
    ws['C4'] = menu_upload.anketa_values[0]
    ws['C5'] = menu_upload.anketa_values[1]
    ws['C6'] = menu_upload.anketa_values[2]
    ws['C7'] = menu_upload.anketa_values[4]
    # 'Результаты проверки по местам работы'
    ws['A9'] = menu_upload.anketa_values[-6]
    ws['C9'] = menu_upload.anketa_values[-5]
    ws['A10'] = menu_upload.anketa_values[-4]
    ws['C10'] = menu_upload.anketa_values[-3]
    ws['A11'] = menu_upload.anketa_values[-2]
    ws['C11'] = menu_upload.anketa_values[-1]
    # проверка паспорт
    ws['C13'] = check_main.response_check[4]
    # 'Проверка статуса самозанятого'
    ws['C14'] = check_main.response_check[0]
    # 'Проверка ИНН'
    ws['C15'] = check_main.response_check[1]
    # 'Проверка по списку кандидатов'
    ws['C21'] = check_main.response_check[2]
    # 'Проверка по списку дисквалифицированных лиц'
    ws['C20'] = check_main.response_check[3]
    # 'Дата проверки'
    ws['C26'] = date.today().strftime('%d.%m.%Y')
    # имя пользователя
    ws['C27'] = getpass.getuser()

    # создаем папку и сохраняем файл с новым именем
    os.makedirs('/home/semenenko/MyProjects/' + menu_upload.anketa_values[2], exist_ok=True)
    wb.save('/home/semenenko/MyProjects/' + menu_upload.anketa_values[2] + '/' + 'Заключение ' +
            menu_upload.anketa_values[2] + '-' + menu_upload.anketa_values[4] + '.xlsx')
    
    db_add()

    mb.showinfo(title="Результат проверки", message="Создана папка и файл заключения. Добавлена запись в БД")
