import os.path
import sqlite3
import shutil
from datetime import date
import os

import openpyxl

# сегодняшняя дата
DATE_TODAY = date.today().strftime('%Y-%m-%d') + ' 00:00:00'
# поля в базе данных
SQL = 'staff, department, full_name, last_name,	birthday, birth_place, country, serie_passport, number_passport, date_given, snils, inn, reg_address, live_address, phone, email, education, check_work_place, check_passport, check_debt, check_bancrupcy, check_bki, check_affilate, check_internet, check_cronos, check_cross, check_rand_info, resume, date_check, officer'
# главный файл кандидатов
MAIN_FILE = r'C:\Users\ubuntu\Documents\Отдел корпоративной защиты\Кандидаты\Кандидаты.xlsm'
# рабочая папка кандидатов
WORK_DIR = r'C:\Users\ubuntu\Documents\Отдел корпоративной защиты\Кандидаты\\'
# место хранения отработанных кандидатов
DEST_DIR = r'C:\Users\ubuntu\Documents\Отдел корпоративной защиты\Персонал\Персонал-2\\'
# база данных где находится реестр и результаты проверки
CONNECT = r'C:\Users\ubuntu\Documents\Отдел корпоративной защиты\Кандидаты\candidates.db'

# функция для записи в БД
def response_db(db, query):
    try:
        with sqlite3.connect(db) as con:
            cur = con.cursor()
            cur.execute(query)
            con.commit()
    except sqlite3.Error as error:
        print('Ошибка', error)       

# Создаем резервную копию книги Exel в', DEST_DIR
shutil.copy(MAIN_FILE, DEST_DIR)
## Создаем резервную копию БД в', DEST_DIR
shutil.copy(CONNECT, DEST_DIR)

# Открываем книгу по адресу:', MAIN_FILE, 'для чтения и записи данных'
wb = openpyxl.load_workbook(MAIN_FILE, keep_vba=True)
# Берем первый лист
ws = wb.worksheets[0]
# Идет поиск ячеек с датами согласования сегодня') #(огранbчение 30 тыс. строк)
col_range = ws['K1':'K30000']
for cell in col_range:
    for c in cell:
        # если запись в ячейке равна сегодняшней дате
        if str(c.value) == DATE_TODAY:
            # берем номер строки
            row_num = c.row
            # берем значения из ячеек, которые советствууюет номеру строки с сегодняшней датой
            fio = ws['B'+str(row_num)].value
            birthday = str(ws['C'+str(row_num)].value)
            staff = ws['D'+str(row_num)].value
            checks = ws['E'+str(row_num)].value
            recruter = ws['F'+str(row_num)].value
            date_in = str(ws['G'+str(row_num)].value)
            officer = ws['H'+str(row_num)].value
            date_out = str(ws['I'+str(row_num)].value)
            result = ws['J'+str(row_num)].value
            fin_date = str(ws['K'+str(row_num)].value)
            url = ws['L'+str(row_num)].value
            #Перебираем каталоги в исходной папке
            for dirs, subdirs, files in os.walk(WORK_DIR):
                for subdir in subdirs:
                    #если имя папки такое же как и значение в ячейке фамилия
                    if subdir.lower().rstrip() == fio.lower().rstrip():
                        # ищем в папке файлы Заключение
                        for file in os.listdir(f"{WORK_DIR[0:-1]+subdir}\\"):
                            if file.startswith("Заключение"):
                                # открываем заключение
                                wbz = openpyxl.load_workbook(os.path.join(WORK_DIR[0:-2], subdir, file), keep_vba=True)
                                
                                # проверяем количество листов если больше 1, то открываем второй лист с анкетой
                                if len(wbz.sheetnames) > 1:
                                    ws1 = wbz.worksheets[1]
                                    # проверяем имеются ли данные на листе с анкетой
                                    if str(ws1['K1'].value) == 'ФИО':
                                        # записываем значения
                                        anketa_values = [str(i) for i in [ws1['C3'].value, ws1['D3'].value, ws1['K3'].value, ws1['S3'].value, ws1['L3'].value, ws1['M3'].value, ws1['T3'].value, ws1['P3'].value, ws1['Q3'].value, ws1['R3'].value, ws1['U3'].value, ws1['V3'].value, ws1['N3'].value, ws1['O3'].value, ws1['Y3'].value, ws1['Z3'].value, ws1['X3'].value]]
                                    # если анкета отсутствует берем анкетные данные с листа заключение
                                    else:
                                        ws2 = wbz.worksheets[0]
                                        anketa_values = [str(i) for i in [ws2['C4'].value, ws2['C5'].value, ws2['C6'].value, ws2['C7'].value, ws2['C8'].value, 'None', 'None', ws2['C9'].value, ws2['D9'].value, ws2['E9'].value, 'None', ws2['C10'].value, 'None', 'None', 'None', 'None','None']]

                                # если второй лсит отсутствует берем анкету с листа заключения
                                elif len(wbz.sheetnames) == 1:
                                    ws2 = wbz.worksheets[0]
                                    anketa_values = [str(i) for i in [ws2['C4'].value, ws2['C5'].value, ws2['C6'].value, ws2['C7'].value, ws2['C8'].value, 'None', 'None', ws2['C9'].value, ws2['D9'].value, ws2['E9'].value, 'None', ws2['C10'].value, 'None', 'None', 'None', 'None','None']]

                                # открываем первый лист с заключением
                                ws2 = wbz.worksheets[0]
                                # записываем значения
                                check_values = [str(i) for i in [
                                    str(ws2['C11'].value) + ' - ' + str(ws2['D11'].value) + '; ' + str(ws2['C12'].value) + ' - ' + str(ws2['D12'].value) + '; ' + str (ws2['C13'].value) + ' - ' + str(ws2['D13'].value), 
                                    str(ws2['B14'].value) + ': ' + str(ws2['C14'].value) + '; ' + str(ws2['B15'].value) + ': ' + str(ws2['C15'].value), 
                                    ws2['C16'].value, 
                                    ws2['C17'].value, 
                                    ws2['C18'].value, 
                                    ws2['C19'].value, 
                                    ws2['C20'].value, 
                                    ws2['C21'].value, 
                                    ws2['C22'].value, 
                                    ws2['C23'].value, 
                                    ws2['C24'].value,
                                    ws2['C25'].value, 
                                    ws2['C29'].value]]
                                # создаем переменную с запросом
                                qu = (f"INSERT INTO candidates ({SQL}) VALUES ('{anketa_values[0]}', '{anketa_values[1]}', '{anketa_values[2]}', '{anketa_values[3]}', '{anketa_values[4]}', '{anketa_values[5]}', '{anketa_values[6]}', '{anketa_values[7]}', '{anketa_values[8]}', '{anketa_values[9]}', '{anketa_values[10]}', '{anketa_values[11]}', '{anketa_values[12]}', '{anketa_values[13]}', '{anketa_values[14]}', '{anketa_values[15]}', '{anketa_values[16]}', '{check_values[0]}', '{check_values[3]}', '{check_values[4]}', '{check_values[5]}', '{check_values[6]}', '{check_values[7]}', '{check_values[8]}', '{check_values[1]}', '{check_values[2]}', '{check_values[12]}', '{check_values[9]}', '{check_values[10]}', '{check_values[11]}')")
                                # заносим результаты проверки в БД
                                response = response_db(CONNECT, qu)
                                wbz.close()
                        #разбираем посимвольно имя папки
                        letter = [i for i in fio]
                        #создаем ссылку для целевой деректории с первым именем буквы по алфавиту, добавляем к имени уникальный ID
                        hlink = DEST_DIR+'\\'+letter[0]+'\\'+subdir+' - '+str(ws['A'+str(row_num)].value)
                        # Создана гиперссылка:', hlink
                        #переносим папку из исходной в целевую папку
                        # shutil.move(WORK_DIR+subdir, hlink)
                        #добавляем в файл книги гиперссылку, куда помещена папка
                        ws['L'+str(row_num)].hyperlink = hlink
                    else:
                        continue
            # передаем данные в таблицу реестр БД
            query = (f"INSERT INTO reestr (fio, birthday, staff, checks, recruter, date_in, officer, date_out, result, final_date, url) VALUES ('{fio}', '{birthday[8:10]}.{birthday[5:7]}.{birthday[0:4]}', '{staff}', '{checks}', '{recruter}', '{date_in}', '{officer}', '{date_out}', '{result}', '{fin_date}', '{url}')")
            # resp = response_db(CONNECT, query)
        else:
            continue                
# print('Cохраняем книгу Excel')
wb.save(MAIN_FILE)
